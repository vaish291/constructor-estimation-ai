import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1E293B")
SECTION_FONT = Font(bold=True, size=12, color="1E293B")


def _style_header_row(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 3, 12), 45)


def export_boq_to_excel(project_name, metrics, boq, optimization=None):
    """Builds an in-memory .xlsx workbook containing the full BOQ."""
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"AI Construction Estimate - {project_name}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    rows = [
        ("Built-up Area (sq.m)", metrics.get("built_up_area_sqm")),
        ("Perimeter (m)", metrics.get("perimeter_m")),
        ("Room Count", metrics.get("room_count")),
        ("Wall Segments Detected", metrics.get("wall_count")),
        ("Door Count", metrics.get("door_count")),
        ("Window Count", metrics.get("window_count")),
        ("Slab Area (sq.m)", boq.get("slab_area_sqm")),
        ("Paint Area (sq.m)", boq.get("paint_area_sqm")),
        ("Tile Area (sq.m)", boq.get("tile_area_sqm")),
        ("", ""),
        ("Material Cost (Rs.)", boq.get("material_cost")),
        ("Labour Cost (Rs.)", boq.get("labour_cost")),
        ("Wastage Amount (Rs.)", boq.get("wastage_amount")),
        (f"GST @ {boq.get('gst_pct')}% (Rs.)", boq.get("gst_amount")),
        (f"Contractor Profit @ {boq.get('profit_pct')}% (Rs.)", boq.get("profit_amount")),
        ("Grand Total AI Estimate (Rs.)", boq.get("total_ai_cost")),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        r += 1
    _autofit(ws)

    # --- Material BOQ sheet ---
    ws2 = wb.create_sheet("Material BOQ")
    headers = ["Item", "Base Qty", "Wastage %", "Total Qty", "Unit", "Rate (Rs.)", "Rate Source", "Total Cost (Rs.)"]
    ws2.append(headers)
    _style_header_row(ws2, 1, len(headers))
    for item in boq.get("material_boq", []):
        ws2.append([
            item.get("item"), item.get("base_quantity"), item.get("wastage_pct"),
            item.get("quantity"), item.get("unit"), item.get("rate"),
            item.get("rate_source"), item.get("total_cost")
        ])
    _autofit(ws2)

    # --- Labour BOQ sheet ---
    ws3 = wb.create_sheet("Labour BOQ")
    headers3 = ["Category", "Days", "Rate/Day (Rs.)", "Total Cost (Rs.)"]
    ws3.append(headers3)
    _style_header_row(ws3, 1, len(headers3))
    for item in boq.get("labour_boq", []):
        ws3.append([item.get("category"), item.get("days"), item.get("rate"), item.get("total_cost")])
    _autofit(ws3)

    # --- AI Optimization sheet ---
    if optimization and optimization.get("suggestions"):
        ws4 = wb.create_sheet("AI Cost Optimization")
        headers4 = ["Original Item", "Suggested Alternative", "Current Cost (Rs.)", "Optimized Cost (Rs.)", "Estimated Saving (Rs.)", "Note"]
        ws4.append(headers4)
        _style_header_row(ws4, 1, len(headers4))
        for s in optimization["suggestions"]:
            ws4.append([
                s.get("original_item"), s.get("suggested_alternative"),
                s.get("current_cost"), s.get("optimized_cost"),
                s.get("estimated_saving"), s.get("note")
            ])
        ws4.append([])
        ws4.append(["Total Estimated Saving (Rs.)", "", "", "", optimization.get("total_estimated_saving")])
        _autofit(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_rates_excel(file_stream):
    """
    Parses an uploaded .xlsx workbook for bulk rate import.

    Expected sheets (optional, either or both may be present):
      - "Material Rates" with columns: location_id, material_name, unit,
        rate_per_unit, rate_source, wastage_pct
      - "Labour Rates" with columns: schedule_type, category, rate_per_day
    """
    wb = load_workbook(file_stream, data_only=True)
    materials = []
    labour = []

    if "Material Rates" in wb.sheetnames:
        ws = wb["Material Rates"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            location_id, material_name, unit, rate_per_unit = row[0], row[1], row[2], row[3]
            rate_source = row[4] if len(row) > 4 and row[4] else "Central (CPWD)"
            wastage_pct = row[5] if len(row) > 5 and row[5] is not None else 5.0
            materials.append({
                "location_id": int(location_id),
                "material_name": str(material_name),
                "unit": str(unit),
                "rate_per_unit": float(rate_per_unit),
                "rate_source": str(rate_source),
                "wastage_pct": float(wastage_pct),
            })

    if "Labour Rates" in wb.sheetnames:
        ws = wb["Labour Rates"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            schedule_type, category, rate_per_day = row[0], row[1], row[2]
            labour.append({
                "schedule_type": str(schedule_type),
                "category": str(category),
                "rate_per_day": float(rate_per_day),
            })

    return {"materials": materials, "labour": labour}


def build_rate_import_template():
    """Generates a blank importable template workbook for admins to fill in."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Material Rates"
    headers = ["location_id", "material_name", "unit", "rate_per_unit", "rate_source", "wastage_pct"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.append([1, "Cement", "Bag", 380.0, "Central (CPWD)", 3.0])
    _autofit(ws)

    ws2 = wb.create_sheet("Labour Rates")
    headers2 = ["schedule_type", "category", "rate_per_day"]
    ws2.append(headers2)
    _style_header_row(ws2, 1, len(headers2))
    ws2.append(["Central (CPWD)", "Mason", 850.0])
    _autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
